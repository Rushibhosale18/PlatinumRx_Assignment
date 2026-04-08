import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from datetime import datetime, timedelta

# Create a new workbook and select active worksheet
wb = openpyxl.Workbook()

# --- SHEET 1: TICKETS ---
ws_ticket = wb.active
ws_ticket.title = "ticket"

# Create some impressive dummy data
tickets_data = [
    ['ticket_id', 'created_at', 'closed_at', 'outlet_id', 'cms_id'],
    ['isu-sjd-457', '2021-08-19 16:45:43', '2021-08-22 12:33:32', 'wrqy-juv-978', 'vew-iuvd-12'],
    ['qer-fal-092', '2021-08-21 11:09:22', '2021-08-21 17:13:45', '8woh-k3u-23b', 'qwe-asd-99'],
    # Adding more data, some with "same day" and "same hour"
    ['tk-aaa-111', '2021-08-22 10:15:00', '2021-08-22 18:30:00', 'wrqy-juv-978', 'cms-101'], # same day, diff hour
    ['tk-bbb-222', '2021-08-23 09:10:00', '2021-08-23 09:45:00', '8woh-k3u-23b', 'cms-102'], # same day, same hour
    ['tk-ccc-333', '2021-08-24 14:20:00', '2021-08-25 10:00:00', 'wrqy-juv-978', 'cms-103'], # diff day
    ['tk-ddd-444', '2021-08-25 11:05:00', '2021-08-25 11:55:00', 'wrqy-juv-978', 'cms-104'], # same day, same hour
    ['tk-eee-555', '2021-08-26 16:45:43', '2021-08-26 21:33:32', 'outlet-alpha', 'cms-105'], # same day, diff hour
    ['tk-fff-666', '2021-08-27 08:30:00', '2021-08-27 08:59:00', 'outlet-alpha', 'cms-106'], # same day, same hour
    ['tk-ggg-777', '2021-08-28 12:00:00', '2021-08-30 12:00:00', '8woh-k3u-23b', 'cms-107'], # diff day
]

for row in tickets_data:
    ws_ticket.append(row)

# Add Helper columns for analysis
ws_ticket['F1'] = 'same_day_flag'
ws_ticket['G1'] = 'same_hour_flag'
for row_num in range(2, len(tickets_data) + 1):
    # Same Day: Compare Date part
    ws_ticket[f'F{row_num}'] = f'=IF(INT(B{row_num})=INT(C{row_num}), "Yes", "No")'
    # Same Hour: Compare Date AND Hour part
    ws_ticket[f'G{row_num}'] = f'=IF(AND(INT(B{row_num})=INT(C{row_num}), HOUR(B{row_num})=HOUR(C{row_num})), "Yes", "No")'

# --- SHEET 2: FEEDBACKS ---
ws_feedback = wb.create_sheet(title="feedbacks")

feedbacks_data = [
    ['cms_id', 'feedback_at', 'feedback_rating', 'ticket_created_at'],
    ['vew-iuvd-12', '2021-08-21 13:26:48', 3, ""],
    ['qwe-asd-99', '2021-08-22 09:12:34', 5, ""],
    ['cms-101', '2021-08-23 10:00:00', 4, ""],
    ['cms-102', '2021-08-24 11:00:00', 5, ""],
    ['cms-104', '2021-08-26 09:00:00', 2, ""],
    ['cms-106', '2021-08-28 14:00:00', 4, ""],
]

for row in feedbacks_data:
    ws_feedback.append(row)

# Add INDEX-MATCH formula dynamically
for row_num in range(2, len(feedbacks_data) + 1):
    ws_feedback[f'D{row_num}'] = f'=IFERROR(INDEX(ticket!$B$2:$B$100, MATCH(A{row_num}, ticket!$E$2:$E$100, 0)), "Not Found")'

# --- SHEET 3: DASHBOARD / ANALYSIS ---
ws_analysis = wb.create_sheet(title="Analysis Summary", index=0)

# Title
ws_analysis.merge_cells('A1:E2')
ws_analysis['A1'] = 'Ticket Resolution Time Analysis'
ws_analysis['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_analysis['A1'].fill = PatternFill('solid', fgColor='4F81BD')
ws_analysis['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws_analysis.append([])

# Headers for Analysis
headers = ['Outlet ID', 'Total Tickets', 'Closed Same Day', 'Closed Same Hour', 'Same Hour %']
ws_analysis.append(headers)

# Formatting headers
for col in range(1, 6):
    cell = ws_analysis.cell(row=4, column=col)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='366092')
    cell.alignment = Alignment(horizontal='center')

# List of unique outlets
outlets = ['wrqy-juv-978', '8woh-k3u-23b', 'outlet-alpha']

row_start = 5
for i, outlet in enumerate(outlets, start=row_start):
    ws_analysis[f'A{i}'] = outlet
    # Total tickets
    ws_analysis[f'B{i}'] = f'=COUNTIF(ticket!D:D, "{outlet}")'
    # Closed same day
    ws_analysis[f'C{i}'] = f'=COUNTIFS(ticket!D:D, "{outlet}", ticket!F:F, "Yes")'
    # Closed same hour
    ws_analysis[f'D{i}'] = f'=COUNTIFS(ticket!D:D, "{outlet}", ticket!G:G, "Yes")'
    # Percentage
    ws_analysis[f'E{i}'] = f'=IF(B{i}=0, 0, D{i}/B{i})'
    ws_analysis[f'E{i}'].number_format = '0.00%'

# Add styling to tables
def style_table(sheet, max_col):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.row == 1 and sheet.title != "Analysis Summary": # Header style
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='4F81BD')
            else:
                cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Apply styles
style_table(ws_ticket, ws_ticket.max_column)
style_table(ws_feedback, ws_feedback.max_column)
style_table(ws_analysis, 5) # For summary sheet

# Set Column Widths for better viewing
for sheet in [ws_ticket, ws_feedback, ws_analysis]:
    for col in sheet.iter_cols():
        max_length = 0
        if not col or not col[0]:
            continue
        column_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        sheet.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
output_path = "c:/Users/Rushi/OneDrive/Desktop/PlatinumRx_Assignment/Data_Analyst_Assignment/Spreadsheets/Ticket_Analysis.xlsx"
wb.save(output_path)
print("Advanced Excel file created successfully at " + output_path)
